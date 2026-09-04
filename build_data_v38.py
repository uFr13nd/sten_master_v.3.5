from openpyxl import load_workbook
from pathlib import Path
from datetime import datetime, date
import json, math

SRC=Path('/mnt/data/sten_v37_current.xlsx')
OUT=Path('/mnt/data/sten-master-web-v38-release/data.js')
wb=load_workbook(SRC,data_only=True,read_only=False)

def val(ws, cell): return ws[cell].value

def n(v):
    if v is None or v=='': return 0.0
    if isinstance(v,(int,float)):
        return float(v) if math.isfinite(float(v)) else 0.0
    s=str(v).replace('\xa0','').replace(' ','').replace(',','.')
    import re
    s=re.sub(r'[^0-9.\-]','',s)
    try:return float(s)
    except:return 0.0

def s(v):
    if v is None:return ''
    if isinstance(v,(datetime,date)):return v.isoformat(sep=' ')
    return str(v).strip()

# refs
works=[]
ws=wb['Работы']
for row in ws.iter_rows(min_row=2,max_row=ws.max_row,min_col=1,max_col=10,values_only=True):
    if not s(row[0]) or not s(row[1]): continue
    works.append(dict(code=s(row[0]),name=s(row[1]),unit=s(row[2]),price=n(row[3]),previousPrice=row[4],changedAt=s(row[5]),technicalGroup=s(row[7]),usage=s(row[8]),comment=s(row[9])))

materials=[]
ws=wb['Материалы']
for row in ws.iter_rows(min_row=2,max_row=ws.max_row,min_col=1,max_col=19,values_only=True):
    if not s(row[0]) or not s(row[1]): continue
    materials.append(dict(
        code=s(row[0]),name=s(row[1]),unit=s(row[2]),price=n(row[3]),previousPrice=row[4],changedAt=s(row[5]),technicalGroup=s(row[7]),usage=s(row[8]),comment=s(row[9]),purchaseGroup=s(row[10]) or 'Обычные материалы',
        dimensions=s(row[11]),densitySpec=n(row[12]) if row[12] not in (None,'') else None,strength=s(row[13]),characteristicsSource=s(row[14]),verificationStatus=s(row[15]),unitComment=s(row[16]),soundInsulation=s(row[17]),calculatorStatus=s(row[18])
    ))
price={x['code']:x['price'] for x in works+materials}

def parse_constructions(sheet,cat):
    ws=wb[sheet]
    rows=list(ws.iter_rows(min_row=1,max_row=min(ws.max_row,300),min_col=1,max_col=17,values_only=True))
    starts=[]
    for i,row in enumerate(rows[1:],start=1):
        if s(row[15])=='Конструкция' and s(row[0]): starts.append(i)
    out=[]
    for k,start in enumerate(starts):
        end=starts[k+1]-1 if k+1<len(starts) else min(len(rows)-1,start+40)
        top=rows[start]
        name=s(top[0])
        if 'полистирол' in name.lower(): continue
        if cat=='part' and s(top[9]).lower().startswith('не проходит'): continue
        item=dict(name=name,city=s(top[1]),thick=n(top[2]),thickFin=n(top[3]),density=n(top[4]),L=n(top[5]),W=n(top[6]),H=n(top[7]),sheetPrice=n(top[8]),reference=s(top[9]),workCostSheet=n(top[12]),materialCostSheet=n(top[13]),sourceRow=start+1,components=[])
        for rix in range(start+1,end+1):
            row=rows[rix]
            typ=s(row[15]); code=s(row[14])
            if not code or typ not in ('Работа','Материал'): continue
            item['components'].append(dict(code=code,type=typ,name=s(row[2]),price=n(price.get(code)),sheetContribution=n(row[7]),purchaseGroup=s(row[16]),note=s(row[9]),sourceRow=rix+1,normCells={'D':row[3],'E':row[4],'F':row[5]}))
        out.append(item)
    return out

constructions={
    'ext':parse_constructions('База_Наружные','ext'),
    'inter':parse_constructions('База_Межквартирные','inter'),
    'part':parse_constructions('База_Перегородки','part'),
}

# finish
ws=wb['База_Отделка']
def finish_def(typ,subtype,rownums,extra,rate_row):
    comps=[]
    for r in rownums:
        code=s(ws.cell(r,15).value); ctype=s(ws.cell(r,16).value)
        if not code or ctype not in ('Работа','Материал'):continue
        comps.append(dict(code=code,type=ctype,name=s(ws.cell(r,1).value),price=n(price.get(code)),sheetContribution=n(ws.cell(r,4).value),sourceRow=r,purchaseGroup=s(ws.cell(r,17).value)))
    comps.append(dict(code=extra,type='Работа',name='Электрика',price=n(price.get(extra)),sheetContribution=n(price.get(extra)),sourceRow=None,purchaseGroup=''))
    return dict(type=typ,subtype=subtype,sheetRate=n(ws.cell(rate_row,2).value),components=comps)
finish=[dict(type='Без отделки',subtype='Кирпич/блок',sheetRate=0,components=[]),dict(type='Без отделки',subtype='ГВЛВ',sheetRate=0,components=[])]
finish += [
 finish_def('Предчистовая','Кирпич/блок',[8,9,10,11,12,15,16,17,18,19,20,21,22],'W-010',6),
 finish_def('Предчистовая','ГВЛВ',[26,27,28,29,30,31,34,35,36,37,38,39,40,41],'W-009',24),
 finish_def('Чистовая','Кирпич/блок',list(range(46,62)),'W-010',44),
 finish_def('Чистовая','ГВЛВ',list(range(65,82)),'W-009',63),
]

# delivery
ws=wb['База_Доставка']
production=[s(ws.cell(1,c).value) for c in range(2,8)]
dist={}
for r in range(2,11): dist[s(ws.cell(r,1).value)]=[n(ws.cell(r,c).value) for c in range(2,8)]
delivery=dict(productionCities=production,distances=dist,rateKm=n(ws['C13'].value),truckVol=n(ws['C14'].value),truckKg=n(ws['C15'].value))

# defaults
ws=wb['ГЛАВНАЯ СТРАНИЦА']
def sel(mat,fin): return {'material':s(ws[mat].value),'finish':s(ws[fin].value) or 'Без отделки'}
defaults=dict(city=s(ws['D5'].value),saleArea=n(ws['D6'].value),totalArea=n(ws['D7'].value),perimExt=n(ws['D8'].value),perimInter=n(ws['D9'].value),perimPart=n(ws['D10'].value),floorH=n(ws['D11'].value),floors=n(ws['D12'].value),salePrice=n(ws['D13'].value),project={'ext':sel('G8','I8'),'inter':sel('G9','I9'),'part':sel('G10','I10')},scenarios=[
 {'ext':sel('D17','D18'),'inter':sel('D41','D42'),'part':sel('D67','D68')},
 {'ext':sel('G17','G18'),'inter':sel('G41','G42'),'part':sel('G67','G68')},
 {'ext':sel('J17','J18'),'inter':sel('J41','J42'),'part':sel('J67','J68')},
])

# baseline current spreadsheet locations
baseline={'ext':[],'inter':[],'part':[],'summary':[]}
for col in ['D','G','J']:
    get=lambda r:n(ws[f'{col}{r}'].value)
    baseline['ext'].append(dict(thick=get(21),thickFin=get(22),footprint=get(23),density=get(24),quantity=get(26),massFloor=get(27),unitPrice=get(28),costFloor=get(29),costObj=get(30),trucks=get(31),delivery=get(32),finishAreaFloor=get(33),finishCost=get(34),total=get(35),workUnit=get(37),materialUnit=get(38)))
    baseline['inter'].append(dict(thick=get(45),thickFin=get(46),footprint=get(47),density=get(48),quantity=get(50),massFloor=get(51),workUnit=get(52),materialUnit=get(53),unitPrice=get(54),costFloor=get(55),costObj=get(56),trucks=get(57),delivery=get(58),finishAreaFloor=get(59),finishRate=get(60),finishCost=get(61),total=get(62)))
    baseline['part'].append(dict(thick=get(71),thickFin=get(72),footprint=get(73),density=get(74),areaFloor=get(75),quantity=get(76),massFloor=get(77),unitPrice=get(78),costFloor=get(79),costObj=get(80),trucks=get(81),deliveryVolFloor=get(82),trucksFloor=get(83),delivery=get(84),finishAreaFloor=get(85),finishRate=get(86),finishCost=get(87),total=get(88),workUnit=get(90),materialUnit=get(91)))
    baseline['summary'].append(dict(deltaFootprint=get(95),deltaSaleFloor=get(96),deltaSaleObj=get(97),deltaSaleValue=get(98)))
for i,col in enumerate(['D','G','J']):
    get=lambda r:n(ws[f'{col}{r}'].value)
    valcol={'D':'E','G':'H','J':'K'}[col]
    getv=lambda r:n(ws[f'{valcol}{r}'].value)
    baseline['summary'][i].update(changedExt=getv(100),changedInter=getv(101),changedPart=getv(102),changedTotal=getv(103),changedPerTotalArea=getv(104),changedPerSaleArea=getv(105))

payload=dict(ok=True,version='3.8',apiVersion='3.8-fallback',timestamp=datetime.now().isoformat(),source='СТЕН_МАСТЕР_v.3.7',defaults=defaults,baseline=baseline,works=works,materials=materials,constructions=constructions,finish=finish,delivery=delivery,constants={'extOpeningPct':0.15,'interOpeningPct':0.07,'partOpeningPct':0.10,'unitWasteFactor':1.07},certificates=[])

def default(o):
    if isinstance(o,(datetime,date)):return o.isoformat()
    return str(o)
js='/* Auto-generated fallback snapshot from СТЕН_МАСТЕР v3.7 for WEB v3.8. */\n(function(root,factory){\n  const data=factory();\n  if(typeof module!=="undefined"&&module.exports) module.exports=data; else root.STEN_FALLBACK_DATA=data;\n})(typeof globalThis!=="undefined"?globalThis:this,function(){ return '+json.dumps(payload,ensure_ascii=False,separators=(',',':'),default=default)+'; });\n'
OUT.write_text(js,encoding='utf-8')
print('wrote',OUT,'materials',len(materials),'constructions',{k:len(v) for k,v in constructions.items()})
